#Finish output (Graphs)

import requests

import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, PieChart, Series, Reference
from openpyxl.chart.series import DataPoint

import warnings
warnings.filterwarnings('ignore')

# r = requests.get(f"https://prices.azure.com/api/retail/prices?$filter=armRegionName eq 'southafricanorth' and serviceName eq 'Virtual Machines' and priceType eq 'Consumption'")
# response = r.json()

# for i in range(len(response['Items'])):
#     print(response['Items'][i]['armSkuName'], i)
######################################################################################################################################################
# importing data
######################################################################################################################################################
print("Importing data\n ...................................")
# input = pd.read_excel('Bryte costing.xlsx', sheet_name = "Sheet4")
input = pd.read_excel('Templete.xlsx')
VMs = pd.read_excel('IDs.xlsx', sheet_name = 'VMs')
Disks = pd.read_excel('IDs.xlsx', sheet_name = 'Disks')
######################################################################################################################################################
# Getting best Azure VM for each input VM
######################################################################################################################################################
print("Searching for best VMs\n ...................................'")
input['ratio'] = input['Size MB']/1024/input['CPUs']

VMdf = pd.DataFrame()

VMdf['vm'] = input['VM']
VMdf['OS according to the VMware Tools'] = input['OS according to the VMware Tools']

for i in range(len(input)):
  if pd.isna(input.loc[i,'ratio']):
      VMdf.loc[i,'Series'] = np.nan
  elif input.loc[i,'ratio'] <= 2:
      VMdf.loc[i,'Series'] = 'F'
  elif input.loc[i,'ratio'] < 8:
      VMdf.loc[i,'Series']  = 'D'
  else:
      VMdf.loc[i,'Series']  = 'E'

for i in range(len(input)):
  temp = VMs[VMs['Name'].str.split(pat='_',expand=True)[1].str[0] == VMdf['Series'].iloc[i]]
  VMdf.loc[i,'azureVM'] = temp[temp['MemoryInMb'] >= input['Size MB'].iloc[i]]['Name'].iloc[0]
  if input['CPUs'].iloc[i] > VMs[VMs['Name'] == VMdf['azureVM'].iloc[i]]['NumberOfCores'].iloc[0]:
    VMdf.loc[i,'azureVM'] = temp[temp['MemoryInMb'] >= input['Size MB'].iloc[i]]['Name'].iloc[1]
######################################################################################################################################################
# Identifying Licenses needed for each VM
######################################################################################################################################################
print("Determining Licenses\n ...................................'")

# VMdf = pd.merge(VMdf[['vm',	'OS according to the VMware Tools',	'Series',	'azureVM']], VM_IDs[['MaxDataDiskCount', 'MemoryInMb', 'Name', 'NumberOfCores','OsDiskSizeInMb', 'ResourceDiskSizeInMb', 'Windows License','Windows SQL License', 'PAYG/1Y/3Y ID','Red Hat Enterprise Linux License']], how = "left", left_on='azureVM', right_on='Name')

licenses = {
    
    'Windows': { 'codes': ['windows','win'],
      #SQL otherwise normal Windows OS
      'Type': {'SQL': ['sql']}},

    'Linux': { 'codes': ['linux','red','hat','redhat','cent','ubuntu'],
                      
      'Type': {'Ubuntu': ['ubuntu'], 
               'Red Hat': ['linux','red','hat','redhat'], 
               'SQL': ['sql','oracle'],
               'SUSE': ['suse']}} 
}

VMdf['PAYG'] = 0
VMdf['1Year'] = 0
VMdf['3Year'] = 0
VMdf['OS'] = 0
VMdf['OS Type'] = 0
VMdf['Windows SQL'] = 0

for i in range(len(VMdf)):
  if  pd.notna(VMdf.loc[i,'OS according to the VMware Tools']):

    #Base price API call ref
    VMdf.loc[i,'PAYG'] = VMs[VMs['Name'] == VMdf['azureVM'].iloc[i]]['PAYG'].iloc[0]
    VMdf.loc[i,'3Year'] = VMs[VMs['Name'] == VMdf['azureVM'].iloc[i]]['3Year'].iloc[0]
    VMdf.loc[i,'1Year'] = VMs[VMs['Name'] == VMdf['azureVM'].iloc[i]]['1Year'].iloc[0]

    #Check if VM is Windows
    if any(code in VMdf.loc[i,'OS according to the VMware Tools'].lower() for code in licenses['Windows']['codes']):
      VMdf.loc[i,'OS'] = VMs[VMs['Name'] == VMdf['azureVM'].iloc[i]]['Windows OS'].iloc[0]
      VMdf.loc[i,'OS Type'] = 'Windows'
      if any(code in VMdf.loc[i,'OS according to the VMware Tools'].lower() for code in licenses['Windows']['Type']['SQL']) or any(code in VMdf.loc[i,'vm'].lower() for code in licenses['Windows']['Type']['SQL']):
        VMdf.loc[i,'Windows SQL'] = VMs[VMs['Name'] == VMdf['azureVM'].iloc[i]]['Windows SQL'].iloc[0]
    #Check if VM is Linux
    elif any(code in VMdf.loc[i,'OS according to the VMware Tools'].lower() for code in licenses['Linux']['codes']):
        # if any(code in VMdf.loc[i,'OS according to the VMware Tools'].lower() for code in licenses['Linux']['Type']['Red Hat']):
        VMdf.loc[i,'OS'] = VMs[VMs['Name'] == VMdf['azureVM'].iloc[i]]['LinuxOS'].iloc[0]
        VMdf.loc[i,'OS Type'] = 'Linux'
        # elif any(code in VMdf.loc[i,'OS according to the VMware Tools'].lower() for code in licenses['Linux']['Type']['SUSE']):
        #   VMdf.loc[i,'SUSE Linux Enterprise'] = ref[ref['Name'] == VMdf['azureVM'].iloc[i]]['SUSE Linux Enterprise'].iloc[0]
######################################################################################################################################################
# Getting best Azure Disk for each input Disk
######################################################################################################################################################
print("Searching for best Disks\n ...................................")
disks = []
priority = []
colNames = [x for x in input.columns if 'Disk' in x]

# VM	OS according to the VMware Tools

for col in colNames:
  for i in range(len(input)):
    if pd.notna(input.loc[i,col]):
      disks.append(input.loc[i,col])
      if 'sql' in input.loc[i,'VM'].lower():
        priority.append(1)
      else:
        priority.append(0)

disksDF = pd.DataFrame({'MemoryInGb':disks, 'Premium': priority})

for i in range(len(disksDF)):

  #Check for premium
  if disksDF.loc[i,'Premium'] == 0:
    temp = Disks[Disks['Name'].str.contains(pat = 'E')]
    disksDF.loc[i,'azureDisk'] = temp[temp['MemoryInGb'] >= disksDF['MemoryInGb'].iloc[i]]['Name'].iloc[0]
  else:
    temp = Disks[Disks['Name'].str.contains(pat = 'P')]
    disksDF.loc[i,'azureDisk'] = temp[temp['MemoryInGb'] >= disksDF['MemoryInGb'].iloc[i]]['Name'].iloc[0]

for i in range(len(disksDF)):
  disksDF.loc[i,'Price'] = Disks[Disks['Name'] == disksDF['azureDisk'].iloc[i]]['Pricing'].iloc[0]
######################################################################################################################################################
# Grouping VM PAYG AHB
######################################################################################################################################################
print('Grouping data\n ...................................')

dfPAYG = pd.DataFrame(columns=['VM', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs'])

for i in range(len(VMdf)):

    vmName = VMdf['azureVM'].iloc[i]
    os = VMdf['OS Type'].iloc[i]
    pricePerVM = VMdf['PAYG'].iloc[i]

    exist = False

    x = 0
    y = 0

    while x < len(dfPAYG):
        if (vmName == dfPAYG['VM'].iloc[x]) and (os == dfPAYG['OS'].iloc[x]):  
            exist = True
            y = x
        x = x+1
    if exist == True:
        dfPAYG['Number'].iloc[y] = dfPAYG['Number'].iloc[y] + 1
        dfPAYG['Monthly Cost for VMs'].iloc[y] = dfPAYG['Monthly Cost for VMs'].iloc[y] + pricePerVM
    else:
        newRow = [vmName, os, 1 , pricePerVM, pricePerVM]
        dfPAYG.loc[len(dfPAYG)] = newRow

num = np.sum(dfPAYG['Number'])
tot_price = np.sum(dfPAYG['Monthly Cost for VMs'])
newRow = ['Total', np.nan, num, np.nan, tot_price]
dfPAYG = dfPAYG.sort_values(by = ['VM'])
dfPAYG.loc[len(dfPAYG)] = newRow

dfPAYG = pd.merge(dfPAYG,VMs[['ResourceDiskSizeInMb','NumberOfCores', 'Name']], left_on='VM', right_on='Name', how='left')
dfPAYG = dfPAYG[['VM','ResourceDiskSizeInMb','NumberOfCores', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs']]
######################################################################################################################################################
# Grouping VM PAYG with License
######################################################################################################################################################
dfPAYG_license = pd.DataFrame(columns=['VM', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs', 'SQL Price Per Month', 'OS Price Per Month', 'Total Price'])

for i in range(len(VMdf)):

    vmName = VMdf['azureVM'].iloc[i]
    os = VMdf['OS Type'].iloc[i]
    pricePerVM = VMdf['PAYG'].iloc[i]
    sqlPricing = VMdf['Windows SQL'].iloc[i]
    osPricing = VMdf['OS'].loc[i]

    exist = False

    x = 0
    y = 0

    while x < len(dfPAYG_license):
        # if (vmName == dfPAYG_license['VM'].iloc[x]) and (os == dfPAYG_license['OS'].iloc[x]):  
        if (vmName == dfPAYG_license['VM'].iloc[x]) and (os == dfPAYG_license['OS'].iloc[x]) and (sqlPricing == dfPAYG_license['SQL Price Per Month'].iloc[x]):  
            exist = True
            y = x
        x = x+1
    if exist == True:
        dfPAYG_license['Number'].iloc[y] = dfPAYG_license['Number'].iloc[y] + 1
        dfPAYG_license['Monthly Cost for VMs'].iloc[y] = dfPAYG_license['Monthly Cost for VMs'].iloc[y] + pricePerVM
        dfPAYG_license['SQL Price Per Month'].iloc[y] = dfPAYG_license['SQL Price Per Month'].iloc[y] + sqlPricing
        dfPAYG_license['OS Price Per Month'].iloc[y] = dfPAYG_license['OS Price Per Month'].iloc[y] + osPricing
        dfPAYG_license['Total Price'].iloc[y] = dfPAYG_license['Total Price'].iloc[y] + pricePerVM + sqlPricing + osPricing
    else:
        tot_price = pricePerVM + sqlPricing + osPricing
        newRow = [vmName, os, 1 , pricePerVM, pricePerVM, sqlPricing, osPricing, tot_price]
        dfPAYG_license.loc[len(dfPAYG_license)] = newRow

num = np.sum(dfPAYG_license['Number'])
tot_price = np.sum(dfPAYG_license['Total Price'])
newRow = ['Total', np.nan, num , np.nan, np.nan, np.nan, np.nan, tot_price]
dfPAYG_license = dfPAYG_license.sort_values(by = ['VM'])
dfPAYG_license.loc[len(dfPAYG_license)] = newRow

dfPAYG_license = pd.merge(dfPAYG_license,VMs[['ResourceDiskSizeInMb','NumberOfCores', 'Name']], left_on='VM', right_on='Name', how='left')
dfPAYG_license = dfPAYG_license[['VM','ResourceDiskSizeInMb','NumberOfCores', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs', 'SQL Price Per Month', 'OS Price Per Month', 'Total Price']]
######################################################################################################################################################
# Grouping VM 3 Year Reserved with AHB
######################################################################################################################################################
df3Y_res = pd.DataFrame(columns=['VM', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs'])

for i in range(len(VMdf)):

    vmName = VMdf['azureVM'].iloc[i]
    os = VMdf['OS Type'].iloc[i]
    pricePerVM = VMdf['3Year'].iloc[i]

    exist = False

    x = 0
    y = 0

    while x < len(df3Y_res):
        if (vmName == df3Y_res['VM'].iloc[x]) and (os == df3Y_res['OS'].iloc[x]):  
            exist = True
            y = x
        x = x+1
    if exist == True:
        df3Y_res['Number'].iloc[y] = df3Y_res['Number'].iloc[y] + 1
        df3Y_res['Monthly Cost for VMs'].iloc[y] = df3Y_res['Monthly Cost for VMs'].iloc[y] + pricePerVM
    else:
        newRow = [vmName, os, 1 , pricePerVM, pricePerVM]
        df3Y_res.loc[len(df3Y_res)] = newRow

num = np.sum(df3Y_res['Number'])
tot_price = np.sum(df3Y_res['Monthly Cost for VMs'])
newRow = ['Total', np.nan, num, np.nan, tot_price]
df3Y_res = df3Y_res.sort_values(by = ['VM'])
df3Y_res.loc[len(df3Y_res)] = newRow

df3Y_res = pd.merge(df3Y_res,VMs[['ResourceDiskSizeInMb','NumberOfCores', 'Name']], left_on='VM', right_on='Name', how='left')
df3Y_res = df3Y_res[['VM','ResourceDiskSizeInMb','NumberOfCores', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs']]
######################################################################################################################################################
# Grouping VM 1 Year Reserved with AHB
######################################################################################################################################################
df1Y_res = pd.DataFrame(columns=['VM', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs'])

for i in range(len(VMdf)):

    vmName = VMdf['azureVM'].iloc[i]
    os = VMdf['OS Type'].iloc[i]
    pricePerVM = VMdf['1Year'].iloc[i]

    exist = False

    x = 0
    y = 0

    while x < len(df1Y_res):
        if (vmName == df1Y_res['VM'].iloc[x]) and (os == df1Y_res['OS'].iloc[x]):  
            exist = True
            y = x
        x = x+1
    if exist == True:
        df1Y_res['Number'].iloc[y] = df1Y_res['Number'].iloc[y] + 1
        df1Y_res['Monthly Cost for VMs'].iloc[y] = df1Y_res['Monthly Cost for VMs'].iloc[y] + pricePerVM
    else:
        newRow = [vmName, os, 1 , pricePerVM, pricePerVM]
        df1Y_res.loc[len(df1Y_res)] = newRow

num = np.sum(df1Y_res['Number'])
tot_price = np.sum(df1Y_res['Monthly Cost for VMs'])
newRow = ['Total', np.nan, num, np.nan, tot_price]
df1Y_res = df1Y_res.sort_values(by = ['VM'])
df1Y_res.loc[len(df1Y_res)] = newRow

df1Y_res = pd.merge(df1Y_res,VMs[['ResourceDiskSizeInMb','NumberOfCores', 'Name']], left_on='VM', right_on='Name', how='left')
df1Y_res = df1Y_res[['VM','ResourceDiskSizeInMb','NumberOfCores', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs']]

######################################################################################################################################################
# Grouping VM 3 Year Reserved with License
######################################################################################################################################################
df3Y_res_license = pd.DataFrame(columns=['VM', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs', 'SQL Price Per Month', 'OS Price Per Month', 'Total Price'])

for i in range(len(VMdf)):

    vmName = VMdf['azureVM'].iloc[i]
    os = VMdf['OS Type'].iloc[i]
    pricePerVM = VMdf['3Year'].iloc[i]
    sqlPricing = VMdf['Windows SQL'].iloc[i]
    osPricing = VMdf['OS'].loc[i]

    exist = False

    x = 0
    y = 0

    while x < len(df3Y_res_license):
        if (vmName == df3Y_res_license['VM'].iloc[x]) and (os == df3Y_res_license['OS'].iloc[x]) and (sqlPricing == df3Y_res_license['SQL Price Per Month'].iloc[x]):  
            exist = True
            y = x
        x = x+1
    if exist == True:
        df3Y_res_license['Number'].iloc[y] = df3Y_res_license['Number'].iloc[y] + 1
        df3Y_res_license['Monthly Cost for VMs'].iloc[y] = df3Y_res_license['Monthly Cost for VMs'].iloc[y] + pricePerVM
        df3Y_res_license['SQL Price Per Month'].iloc[y] = df3Y_res_license['SQL Price Per Month'].iloc[y] + sqlPricing
        df3Y_res_license['OS Price Per Month'].iloc[y] = df3Y_res_license['OS Price Per Month'].iloc[y] + osPricing
        df3Y_res_license['Total Price'].iloc[y] = df3Y_res_license['Total Price'].iloc[y] + pricePerVM + sqlPricing + osPricing
    else:
        tot_price = pricePerVM + sqlPricing + osPricing
        newRow = [vmName, os, 1 , pricePerVM, pricePerVM, sqlPricing, osPricing, tot_price]
        df3Y_res_license.loc[len(df3Y_res_license)] = newRow

num = np.sum(df3Y_res_license['Number'])
tot_price = np.sum(df3Y_res_license['Total Price'])
newRow = ['Total', np.nan, num , np.nan, np.nan, np.nan, np.nan, tot_price]
df3Y_res_license = df3Y_res_license.sort_values(by = ['VM'])
df3Y_res_license.loc[len(df3Y_res_license)] = newRow

df3Y_res_license = pd.merge(df3Y_res_license,VMs[['ResourceDiskSizeInMb','NumberOfCores', 'Name']], left_on='VM', right_on='Name', how='left')
df3Y_res_license = df3Y_res_license[['VM','ResourceDiskSizeInMb','NumberOfCores', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs', 'SQL Price Per Month', 'OS Price Per Month', 'Total Price']]
######################################################################################################################################################
# Grouping VM 3 Year Reserved with License
######################################################################################################################################################
df1Y_res_license = pd.DataFrame(columns=['VM', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs', 'SQL Price Per Month', 'OS Price Per Month', 'Total Price'])

for i in range(len(VMdf)):

    vmName = VMdf['azureVM'].iloc[i]
    os = VMdf['OS Type'].iloc[i]
    pricePerVM = VMdf['1Year'].iloc[i]
    sqlPricing = VMdf['Windows SQL'].iloc[i]
    osPricing = VMdf['OS'].loc[i]

    exist = False

    x = 0
    y = 0

    while x < len(df1Y_res_license):
        if (vmName == df1Y_res_license['VM'].iloc[x]) and (os == df1Y_res_license['OS'].iloc[x]) and (sqlPricing == df1Y_res_license['SQL Price Per Month'].iloc[x]):  
            exist = True
            y = x
        x = x+1
    if exist == True:
        df1Y_res_license['Number'].iloc[y] = df1Y_res_license['Number'].iloc[y] + 1
        df1Y_res_license['Monthly Cost for VMs'].iloc[y] = df1Y_res_license['Monthly Cost for VMs'].iloc[y] + pricePerVM
        df1Y_res_license['SQL Price Per Month'].iloc[y] = df1Y_res_license['SQL Price Per Month'].iloc[y] + sqlPricing
        df1Y_res_license['OS Price Per Month'].iloc[y] = df1Y_res_license['OS Price Per Month'].iloc[y] + osPricing
        df1Y_res_license['Total Price'].iloc[y] = df1Y_res_license['Total Price'].iloc[y] + pricePerVM + sqlPricing + osPricing
    else:
        tot_price = pricePerVM + sqlPricing + osPricing
        newRow = [vmName, os, 1 , pricePerVM, pricePerVM, sqlPricing, osPricing, tot_price]
        df1Y_res_license.loc[len(df1Y_res_license)] = newRow

num = np.sum(df1Y_res_license['Number'])
tot_price = np.sum(df1Y_res_license['Total Price'])
newRow = ['Total', np.nan, num , np.nan, np.nan, np.nan, np.nan, tot_price]
df1Y_res_license = df1Y_res_license.sort_values(by = ['VM'])
df1Y_res_license.loc[len(df1Y_res_license)] = newRow

df1Y_res_license = pd.merge(df1Y_res_license,VMs[['ResourceDiskSizeInMb','NumberOfCores', 'Name']], left_on='VM', right_on='Name', how='left')
df1Y_res_license = df1Y_res_license[['VM','ResourceDiskSizeInMb','NumberOfCores', 'OS','Number', 'Price Per VM', 'Monthly Cost for VMs', 'SQL Price Per Month', 'OS Price Per Month', 'Total Price']]
######################################################################################################################################################
# Grouping Disks
######################################################################################################################################################
dfDisksOut = pd.DataFrame(columns=['Disk', 'Size','Number', 'Price Per Disk', 'Monthly Cost for Disks'])

for i in range(len(disksDF)):

    name = disksDF['azureDisk'].iloc[i]
    size = disksDF['MemoryInGb'].iloc[i]
    price = disksDF['Price'].iloc[i]

    exist = False

    x = 0
    y = 0

    while x < len(dfDisksOut):
        if (name == dfDisksOut['Disk'].iloc[x]):  
            exist = True
            y = x
        x = x+1
    if exist == True:
        dfDisksOut['Number'].iloc[y] = dfDisksOut['Number'].iloc[y] + 1
        dfDisksOut['Monthly Cost for Disks'].iloc[y] = dfDisksOut['Monthly Cost for Disks'].iloc[y] + price
    else:
        newRow = [name, size, 1, price, price]
        dfDisksOut.loc[len(dfDisksOut)] = newRow

num = np.sum(dfDisksOut['Number'])
tot_price = np.sum(dfDisksOut['Monthly Cost for Disks'])
newRow = ['Total', np.nan, num , np.nan, tot_price]
dfDisksOut = dfDisksOut.sort_values(by = ['Disk'])
dfDisksOut.loc[len(dfDisksOut)] = newRow
######################################################################################################################################################
# Exporting Data
######################################################################################################################################################
print('Exporting data\n ...................................')

wb = Workbook()
ws = wb.active
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws0 = wb.worksheets[0]
ws0.title = 'VMs'
for r in dataframe_to_rows(VMdf, index=False, header=True):
    ws0.append(r)
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws1 = wb.create_sheet()
ws1.title = 'Disks'
for r in dataframe_to_rows(disksDF, index=False, header=True):
    ws1.append(r)
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws2 = wb.create_sheet()
ws2.title = 'VM_PAYG_AHB'
for r in dataframe_to_rows(dfPAYG, index=False, header=True):
    ws2.append(r)
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws3 = wb.create_sheet()
ws3.title = 'VM_PAYG_License'
for r in dataframe_to_rows(dfPAYG_license, index=False, header=True):
    ws3.append(r)
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws7 = wb.create_sheet()
ws7.title = 'VM_1Reserved_AHB'
for r in dataframe_to_rows(df1Y_res, index=False, header=True):
    ws7.append(r)
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws8 = wb.create_sheet()
ws8.title = 'VM_1Reserved_License'
for r in dataframe_to_rows(df1Y_res_license, index=False, header=True):
    ws8.append(r)
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws4 = wb.create_sheet()
ws4.title = 'VM_3Reserved_AHB'
for r in dataframe_to_rows(df3Y_res, index=False, header=True):
    ws4.append(r)
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws5 = wb.create_sheet()
ws5.title = 'VM_3Reserved_License'
for r in dataframe_to_rows(df3Y_res_license, index=False, header=True):
    ws5.append(r)
#-----------------------------------------------------------------------------------------------------------------------------------------------------
ws6 = wb.create_sheet()
ws6.title = 'Disks_Prices'
for r in dataframe_to_rows(dfDisksOut, index=False, header=True):
    ws6.append(r)
######################################################################################################################################################
# Generating Graphs
######################################################################################################################################################
print('Generating Graphs\n ...................................')

#Bar Chart for totals
wsGraph = wb.create_sheet()
wsGraph.title = 'Charts'

# license_cost =  df3Y_res_license.iloc[len(df3Y_res_license)-1,len(df3Y_res_license.columns)-1] - dfPAYG_license.iloc[len(dfPAYG_license)-1,len(dfPAYG_license.columns)-1]
datas = [
    ['Option', '$ Per Month'],
    ['VM_PAYG_AHB', dfPAYG.iloc[len(dfPAYG)-1,len(dfPAYG.columns)-1]],
    ['VM_PAYG_License', dfPAYG_license.iloc[len(dfPAYG_license)-1,len(dfPAYG_license.columns)-1]],
    ['VM_1Reserved_AHB', df1Y_res.iloc[len(df1Y_res)-1,len(df1Y_res.columns)-1]],
    ['VM_1Reserved_License', df1Y_res_license.iloc[len(df1Y_res_license)-1,len(df1Y_res_license.columns)-1]],
    ['VM_3Reserved_AHB', df3Y_res.iloc[len(df3Y_res)-1,len(df3Y_res.columns)-1]],
    ['VM_3Reserved_License', df3Y_res_license.iloc[len(df3Y_res_license)-1,len(df3Y_res_license.columns)-1]],
    ['License Costs', dfPAYG_license.iloc[len(dfPAYG_license)-1,len(dfPAYG_license.columns)-1] - df3Y_res_license.iloc[len(df3Y_res_license)-1,len(df3Y_res_license.columns)-1]],
]
 
for row in datas:
    wsGraph.append(row)

chart = BarChart()
labels = Reference(wsGraph, min_col = 1,
                   min_row = 2, max_row = 8)
data = Reference(wsGraph, min_col = 2,
                   min_row = 1, max_row = 8)
chart.add_data(data, titles_from_data = True)
chart.set_categories(labels)
chart.title = " Total $ per Month"
 
wsGraph.add_chart(chart, "A1")
#-----------------------------------------------------------------------------------------------------------------------------------------------------
#Bar chart for accumalating values

datas = [
    ['Month','VM_PAYG_AHB','VM_1Reserved_AHB','VM_3Reserved_AHB'],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
    [],
]

x = 0
y = 0
z = 0
for i in range(1,37):
    x = x + dfPAYG.iloc[len(dfPAYG)-1,len(dfPAYG.columns)-1]
    y = y + df1Y_res.iloc[len(df1Y_res)-1,len(df1Y_res.columns)-1]
    z = z + df3Y_res.iloc[len(df3Y_res)-1,len(df3Y_res.columns)-1]
    datas[i].append(i)
    datas[i].append(x)
    datas[i].append(y)
    datas[i].append(z)

for row in datas:
    wsGraph.append(row)

c1 = LineChart()
c1.title = "Accumulated Pricing over 3 Years"
c1.style = 13
c1.y_axis.title = '$'
c1.x_axis.title = 'Months'
c1.style = 10
data = Reference(wsGraph, min_col=2, min_row=9, max_col=4, max_row=45)
c1.add_data(data, titles_from_data=True)
wsGraph.add_chart(c1, "A16")
#-----------------------------------------------------------------------------------------------------------------------------------------------------
#Grouped bar chart for accumalating values
rows = [
    ['Month', 'PAYG', '1 Year Reserved', '3 Year Reserved'],
    [1],
    [2],
    [3],
    [4],
    [5],
    [6],
    [7],
    [8],
    [9],
    [10],
    [11],
    [12],
]

x = 0
y = 0
z = 0
for i in range(1,13):
    x = x + dfPAYG.iloc[len(dfPAYG)-1,len(dfPAYG.columns)-1]
    y = y + df1Y_res.iloc[len(df1Y_res)-1,len(df1Y_res.columns)-1]
    z = z + df3Y_res.iloc[len(df3Y_res)-1,len(df3Y_res.columns)-1]
    rows[i].append(x)
    rows[i].append(y)
    rows[i].append(z)

for row in rows:
    wsGraph.append(row)

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Comparison"
chart1.y_axis.title = '$'
chart1.x_axis.title = 'Months'

data = Reference(wsGraph, min_col=2, min_row=46, max_row=58, max_col=4)
cats = Reference(wsGraph, min_col=1, min_row=47, max_row=58)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
wsGraph.add_chart(chart1, "A20")
#-----------------------------------------------------------------------------------------------------------------------------------------------------
# Pie chart for accumalating values

data = [
    ['Pie', 'Sold'],
    ['Apple', 50],
    ['Cherry', 30],
    ['Pumpkin', 10],
    ['Chocolate', 40],
]

for row in data:
    wsGraph.append(row)

pie = PieChart()
labels = Reference(wsGraph, min_col=1, min_row=59, max_row=63)
data = Reference(wsGraph, min_col=2, min_row=59, max_row=63)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Pies sold by category"

wsGraph.add_chart(pie, "H1")

wb.save("Check.xlsx")

#Storage vs Disk prices: Pie Chart
#